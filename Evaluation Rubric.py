
import streamlit as st
import pdfplumber
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import tempfile
import os

def extract_tables_from_pages(pdf_path, page_numbers):
    with pdfplumber.open(pdf_path) as pdf:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evaluation Rubric"

        for page_num in page_numbers:
            if page_num < 1 or page_num > len(pdf.pages):
                st.warning(f"Page {page_num} is out of range and will be skipped.")
                continue

            page = pdf.pages[page_num - 1]
            tables = page.extract_tables()

            if not tables:
                st.info(f"No tables found on page {page_num}")
                continue

            for table in tables:
                for row in table:
                    clean_row = [cell.strip() if cell else "" for cell in row]
                    ws.append(clean_row)
                ws.append([])

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 5, 50)

        return wb

st.set_page_config(page_title="Evaluation Rubric Extractor", layout="centered")
st.title("📊 Evaluation Rubric Extractor from PDF")
st.markdown("Upload a PDF and specify page numbers to extract tables into an Excel file.")

uploaded_pdf = st.file_uploader("Upload PDF file", type=["pdf"])
page_input = st.text_input("Enter page numbers (comma-separated)", "47,48")

if uploaded_pdf and page_input:
    try:
        page_numbers = [int(num.strip()) for num in page_input.split(",") if num.strip().isdigit()]
        if st.button("📥 Extract Tables"):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
                tmp_pdf.write(uploaded_pdf.read())
                tmp_pdf_path = tmp_pdf.name

            workbook = extract_tables_from_pages(tmp_pdf_path, page_numbers)

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
                workbook.save(tmp_excel.name)
                tmp_excel_path = tmp_excel.name

            with open(tmp_excel_path, "rb") as f:
                st.download_button(
                    label="📤 Download Extracted Excel File",
                    data=f,
                    file_name="Evaluation_Rubric.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            os.remove(tmp_pdf_path)
            os.remove(tmp_excel_path)

    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
